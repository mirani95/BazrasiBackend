from datetime import datetime
from io import BytesIO

import jdatetime
import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

blue_fill = PatternFill(start_color="277358", fill_type="solid")
Alignment_CELL = Alignment(horizontal='center', vertical='center', wrap_text=True)
red_font = Font(color="C00000", bold=True)
GREEN_CELL = PatternFill(start_color="00B050", fill_type="solid")
RED_CELL = PatternFill(start_color="FCDFDC", fill_type="solid")
YELLOW_CELL = PatternFill(start_color="FFFF00", fill_type="solid")
ORANGE_CELL = PatternFill(start_color="FFC000", fill_type="solid")
BLUE_CELL = PatternFill(start_color="538DD5", fill_type="solid")
LIGHT_GREEN_CELL = PatternFill(start_color="92D050", fill_type="solid")
VERY_LIGHT_GREEN_CELL = PatternFill(start_color="5AFC56", fill_type="solid")


def shamsi_date(date, in_value=None):
    if in_value:
        sh_date = jdatetime.date.fromgregorian(
            year=date.year,
            month=date.month,
            day=date.day
        )
    else:
        gh_date = jdatetime.date.fromgregorian(
            year=date.year,
            month=date.month,
            day=date.day
        ).strftime('%Y-%m-%d')
        reversed_date = reversed(gh_date.split("-"))
        separate = "-"
        sh_date = separate.join(reversed_date)
    return sh_date


def create_header(worksheet, list, num, row, height=None, width=None, color=None, text_color=None, border_style=None):
    for col_num, option in enumerate(list, num):
        cell = worksheet.cell(row=row, column=col_num, value=option)
        col_letter = get_column_letter(col_num)
        cell.alignment = Alignment_CELL
        if color is not None:
            if color == 'green':
                cell.fill = GREEN_CELL
            elif color == 'orange':
                cell.fill = ORANGE_CELL
            elif color == 'blue':
                cell.fill = BLUE_CELL
            else:
                cell.fill = PatternFill(start_color=color, fill_type="solid")
        else:
            cell.fill = blue_fill
        if text_color is not None:
            cell.font = Font(size=9, bold=True, color=text_color)
        else:
            cell.font = Font(size=9, bold=True, color='D9FFFFFF')
        if height is not None:
            worksheet.row_dimensions[row].height = height
        if width is not None:
            worksheet.column_dimensions[col_letter].width = width
        if border_style is not None:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=border_style),
                right=openpyxl.styles.Side(style=border_style),
                top=openpyxl.styles.Side(style=border_style),
                bottom=openpyxl.styles.Side(style=border_style)
            )


def create_header_freez(worksheet, list, num, row, header_row, height=None, width=None, len_with=None,
                        different_cell=None):
    for col_num, option in enumerate(list, num):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=row, column=col_num, value=option)
        cell.alignment = Alignment_CELL
        cell.fill = blue_fill
        cell.font = Font(size=10, bold=True, color='D9FFFFFF')

        if height is not None:
            worksheet.row_dimensions[row].height = height
            if len(option) > worksheet.column_dimensions[col_letter].width:
                worksheet.column_dimensions[col_letter].width = len(option) + 2
        if width is not None:
            worksheet.column_dimensions[col_letter].width = width
        if len_with is not None:
            if len(option) > worksheet.column_dimensions[col_letter].width:
                worksheet.column_dimensions[col_letter].width = len(option) + 3
        if different_cell is not None:
            if option == different_cell:
                cell.fill = PatternFill(start_color="C00000", fill_type="solid")
        worksheet.freeze_panes = worksheet[f'A{header_row}']
        max_col = worksheet.max_column
        range_str = f'A{header_row - 1}:{get_column_letter(max_col)}{worksheet.max_row}'
        worksheet.auto_filter.ref = range_str


def excel_description(worksheet, row1, description, size=None, color=None, my_color=None, row2=None):
    worksheet[row1] = description
    worksheet[row1].alignment = Alignment_CELL
    if size is not None:
        worksheet[row1].font = Font(size=size)
    if color is not None:
        worksheet[row1].font = red_font
    if my_color is not None:
        worksheet[row1].font = PatternFill(start_color=my_color, fill_type="solid")

    if row2 is not None:
        merge_range = f'{row1}:{row2}'
        worksheet.merge_cells(merge_range)


def create_value(worksheet, list, l, num, border_style=None, m=None, height=None, color=None, width=None,
                 different_cell=None, different_value=None, item_num=None, item_color=None):
    color_dict = {
        'green': GREEN_CELL,
        'yellow': YELLOW_CELL,
        'blue': BLUE_CELL,
        'red': RED_CELL,
        'light_green': LIGHT_GREEN_CELL,
        'very_light_green': VERY_LIGHT_GREEN_CELL
    }

    for item in range(len(list)):
        cell = worksheet.cell(row=l, column=item + num, value=list[item])
        cell.alignment = Alignment_CELL

        if border_style:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=border_style),
                right=openpyxl.styles.Side(style=border_style),
                top=openpyxl.styles.Side(style=border_style),
                bottom=openpyxl.styles.Side(style=border_style)
            )

        value = list[item]
        if isinstance(value, (int, float)) and value != 0:
            cell.number_format = '#,###'
        else:
            cell.value = value

        cell.font = Font(size=10, bold=True)

        if m is not None and m % 2 != 0:
            cell.fill = PatternFill(start_color="D6F6FE", fill_type="solid")

        if height is not None:
            worksheet.row_dimensions[l + 1].height = height

        if item_num is not None and item == item_num:
            if item_color:
                cell.fill = item_color
        elif color in color_dict:
            cell.fill = color_dict[color]

        if different_cell is not None and list[different_cell] == different_value:
            cell.fill = RED_CELL

        if width is not None:
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(item + num)].width = width


def merge_cells(worksheet, l, s, cell1=None, cell2=None, lst=None):
    if lst is not None:
        for col in lst:
            rng = f'{col}{l}:{col}{l + s}'
            worksheet.merge_cells(rng)
            worksheet[col + f'{l}'].alignment = Alignment_CELL
    else:
        for col in range(ord(f'{cell1}'), ord(f'{cell2}') + 1):
            rng = f'{chr(col)}{l}:{chr(col)}{l + s}'
            worksheet.merge_cells(rng)
            worksheet[chr(col) + f'{l}'].alignment = Alignment_CELL


def add_header(worksheet):
    worksheet.oddHeader.center.text = "سامانه رصدیار"
    worksheet.oddHeader.center.size = 14  # تنظیم اندازه فونت
    worksheet.oddHeader.center.font = "Arial,Bold"  # تنظیم فونت و ضخامت

    # همچنین می‌توانید از هدرهای چپ و راست هم استفاده کنید
    # worksheet.oddHeader.right.text = f"تاریخ: {shamsi_now_date}"


def cell_color_changer(worksheet, row, start_index, end_index, custom_color):
    for item in range(start_index, end_index):
        cell = worksheet.cell(row=row, column=item)
        cell.fill = PatternFill(start_color=custom_color, fill_type="solid")


def start_excel():
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.sheet_view.rightToLeft = True
    worksheet.insert_rows(1)
    return workbook, worksheet, output


def close_excel(name):
    workbook, worksheet, output = start_excel()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename="{name}.xlsx"'.encode(
        'utf-8')
    response.write(output.getvalue())
    return response


def add_chart(
        worksheet,
        chart_type,
        data_columns,
        category_column,
        start_row,
        end_row,
        chart_position,
        chart_title,
        x_axis_title,
        y_axis_title,
        chart_width=25,  # عرض نمودار پیش‌فرض (واحد: cm)
        chart_height=15
):
    """
    افزودن نمودار به صفحه اکسل.

    ورودی:
        worksheet (openpyxl.Worksheet): صفحه اکسل.
        chart_type (str): نوع نمودار ("line" یا "bar").
        data_columns (list): لیستی از ستون‌های داده.
        category_column (int): ستون دسته‌بندی‌ها.
        start_row (int): ردیف شروع داده‌ها.
        end_row (int): ردیف پایان داده‌ها.
        chart_position (str): محل قرار گرفتن نمودار.
        chart_title (str): عنوان نمودار.
        x_axis_title (str): عنوان محور X.
        y_axis_title (str): عنوان محور Y.
        chart_width (float): عرض نمودار (واحد: cm).
        chart_height (float): ارتفاع نمودار (واحد: cm).
    """

    if chart_type == 'line':
        chart = LineChart()
        chart.style = 20
    elif chart_type == 'bar':
        chart = BarChart()
    else:
        raise ValueError("chart_type باید 'line' یا 'bar' باشد.")

    chart.title = chart_title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.width = chart_width
    chart.height = chart_height

    categories = Reference(worksheet, min_col=category_column, min_row=start_row, max_row=end_row)
    data = Reference(worksheet, min_col=data_columns, min_row=start_row - 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    for series in chart.series:
        series.graphicalProperties.line.solidFill = "277358"
        series.graphicalProperties.line.width = 30000

    worksheet.add_chart(chart, chart_position)
    # example
    # add_chart(
    #     worksheet=worksheet,
    #     chart_type='line',
    #     data_columns=7,  # ستون وزن وارد شده
    #     category_column=2,  # ستون نام سردخانه‌ها
    #     start_row=7,
    #     end_row=l + 1,
    #     chart_position="A12",
    #     chart_title="نمودار تغییرات وزن در سردخانه‌ها",
    #     x_axis_title="سردخانه‌ها",
    #     y_axis_title="وزن (کیلوگرم)"
    # )


def to_locale_str(a):
    return "{:,}".format(int(a))


def convert_str_to_date(string):
    return datetime.strptime(str(string), '%Y-%m-%d').date()
